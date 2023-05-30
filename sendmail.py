import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl import Workbook

from dotenv import load_dotenv
load_dotenv()
import os

src_filename = "創客研習證明.xlsx"
id = os.getenv('id')
passwd = os.getenv('passwd')

# 連線並登入旗標的郵件伺服器
print("開始執行.")
smtp = smtplib.SMTP('flag.com.tw', 25)
smtp.login(id, passwd)
print("已登入郵件伺服器.")

# 開啟研習證明的登錄試算表檔
wb = load_workbook(src_filename)
ws = wb.active
print("已開啟結業證書名冊檔.")

#     1          2      3      4     5   6   7      
# 參加人姓名	參加人Email	字號	活動名稱	年	月	日
cource_name = f"{ws.cell(2,6).value:02d}/{ws.cell(2,7).value:02d} {ws.cell(2,4).value}"
date_str = f'{ws.cell(2,5).value}{ws.cell(2,6).value:02d}{ws.cell(2,7).value:02d}'

# 針對試算表中活頁簿的每一列 (第一列是標題) 寄送研習證明
for i in range(2, ws.max_row + 1):
  print(f"第 {i - 1} 筆資料：")
  print(ws.cell(i, 1).value, ws.cell(i, 2).value)

  msg = MIMEMultipart('alternative')
  msg['Subject'] = f'旗標科技 [{cource_name}] 創客體驗營研習證明'
  msg['From'] = 'marketing@flag.com.tw'
  msg['To'] = ws.cell(i, 2).value

  html = f"""
  <html>
    <body>
      <p>{ws.cell(i, 1).value} 學員您好</p>
      <p>謝謝您參加 [{cource_name}] 創客體驗贏, 附件是您的研習證明, 期待下次課程再與您相見。</p>
      <p>旗標科技創客團隊</p>
    </body
  </html>
  """

  # 加入 HTML 格式的信件內容
  html_part = MIMEText(html, "html")
  msg.attach(html_part)

  # 讀取並建立包含研習證明 PDF 檔內容的 MIME 物件
  fname = f'{date_str}_{ws.cell(i, 3).value}.pdf'
  print(fname)

  with open(fname, 'rb') as attach:
    attach_part = MIMEBase('application', 'octet-stream')
    attach_part.set_payload(attach.read())

  # 將加入的 PDF 檔內容以 base64 編碼
  encoders.encode_base64(attach_part)

  # 加入表頭註記這是附加檔案, 不是郵件內容
  attach_part.add_header(
    'Content-Disposition',
    f'attachment; filename={fname}'
  )

  # 將附件加入郵件中
  msg.attach(attach_part)
  # 送出郵件
  smtp.send_message(msg)

smtp.quit()
wb.close()
print("結束程式.")
